import requests
from lxml import etree
import re
import pandas as pd
import openpyxl
from datetime import datetime

url = 'https://list1.mysteel.com/market/p-405-----010211-0-0101040401-------1.html'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36',
    'Cookie': '_last_loginuname=goodbaby123; _login_psd=3ba7e75e56a34c01b56c65f10e8901b16; _rememberStatus=true; Hm_lvt_1c4432afacfa2301369a5625795031b8=1649222144; allHistory=%5B%22%E7%94%B5%E9%95%80%E4%B8%9D%22%5D; href=http%3A%2F%2Fguigang.mysteel.com%2F; accessId=5d36a9e0-919c-11e9-903c-ab24dbab411b; pageViewNum=3; _last_ch_r_t=1649397611208; fingerprint=a64ff2e9c89fde087b97042290980c0c; _login_token=3b4df51e8f109f30b5166808f3ce710b; _login_uid=2243467; _login_mid=3057897; _login_ip=183.129.214.218; 3b4df51e8f109f30b5166808f3ce710b=1%3D10%262%3D10%26catalog%3D010102%2C010103; Hm_lpvt_1c4432afacfa2301369a5625795031b8=1649398288'}

response = requests.get(url)
url_tree = etree.HTML(response.text)
url_check_title = url_tree.xpath('//*[@id="articleList"]/ul/li[1]/a/text()')[0]
# 判断是否为更新的数据
if url_check_title.find("（") == -1:
    next_url = url_tree.xpath('//*[@id="articleList"]/ul/li[1]/a/@href')[0]
else:
    next_url = url_tree.xpath('//*[@id="articleList"]/ul/li[2]/a/@href')[0]
detail_response = requests.get(next_url, headers=headers)
detail_url_tree = etree.HTML(detail_response.text)
# 获取数据标题和日期
detail_title = detail_url_tree.xpath('//*[@id="articleContent"]/h1/text()')[0]
detail_data_date = re.search('([0-9]+?月[0-9]+?日)', detail_title).group()
detail_data_month = detail_data_date.split('月')[0] + '月份（吨）'
detail_data_day = detail_data_date.split('月')[1][0] + '号'
detail_tr_list = detail_url_tree.xpath('//*[@id="marketTable"]/tr')
detail_tr_diandusi = detail_tr_list[3]
detail_tr_pensusi = detail_tr_list[5]


writer = pd.ExcelWriter('./每日明细数据/最新数据' + detail_data_date + '.xlsx')
# 电镀丝
data_list_diandusi = []
td_list_diandusi = detail_tr_diandusi.xpath('./td[@align]')
td_list_diandusi = [td_list_diandusi[0], td_list_diandusi[1], td_list_diandusi[4], td_list_diandusi[5]]
single_data_list_diandusi = []
for td in td_list_diandusi:
    data = td.xpath('./text()')[0].replace('\r', '').replace('\n', '').replace('\t', '').replace(' ', '')
    single_data_list_diandusi.append(data)
data_list_diandusi.append(single_data_list_diandusi)

df_diandusi = pd.DataFrame(data_list_diandusi)
df_diandusi.columns = ['钢种', '牌号', '钢厂', '价格']
print(df_diandusi)
df_diandusi.to_excel(writer, sheet_name='电镀丝', index=False)
# df_diandusi.to_excel('./每日明细数据/最新数据' + detail_data_date + '.xlsx', sheet_name='电镀丝', index=False)

# 喷塑丝
data_list_pensusi = []
td_list_pensusi = detail_tr_pensusi.xpath('./td[@align]')
td_list_pensusi = [td_list_pensusi[0], td_list_pensusi[1], td_list_pensusi[4], td_list_pensusi[5]]
single_data_list_pensusi = []
for td in td_list_pensusi:
    data = td.xpath('./text()')[0].replace('\r', '').replace('\n', '').replace('\t', '').replace(' ', '')
    single_data_list_pensusi.append(data)
data_list_pensusi.append(single_data_list_pensusi)

df_pensusi = pd.DataFrame(data_list_pensusi)
df_pensusi.columns = ['钢种', '牌号', '钢厂', '价格']
print(df_pensusi)
df_pensusi.to_excel(writer, sheet_name='喷塑丝', index=False)
# df_pensusi.to_excel('./每日明细数据/最新数据' + detail_data_date + '.xlsx', sheet_name='喷塑丝', index=False)
writer.save()


date = datetime.now()
today = str(date.month) + '月' + str(date.day) + '日'
day = date.day
month = date.month
file_path = '原始数据.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.sheetnames
sh = wb['电镀丝']
sh_ps = wb['喷塑丝']

new_data_file_path = './每日明细数据/最新数据' + today + '.xlsx'
wb_new = openpyxl.load_workbook(new_data_file_path)
sheet_new = wb_new.sheetnames
sh_new = wb_new['电镀丝']
sh_new_ps = wb_new['喷塑丝']


# 获取待填充单元格坐标
def get_input_loc(day, month):
    for j_loc in range(1, 100, 1):
        if sh.cell(2, j_loc).value == str(month) + '月份':
            col_loc = j_loc
    for i_loc in range(1, 100, 1):
        if sh.cell(i_loc, 1).value == str(day) + '号':
            row_loc = i_loc
    return row_loc, col_loc


sh.cell(get_input_loc(day, month)[0], get_input_loc(day, month)[1]).value = int(sh_new.cell(2, 4).value)
sh_ps.cell(get_input_loc(day, month)[0], get_input_loc(day, month)[1]).value = int(sh_new_ps.cell(2, 4).value)

wb.save('./数据更新/' + today + '更新数据.xlsx')
