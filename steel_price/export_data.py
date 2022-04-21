import pandas as pd
import openpyxl
from datetime import datetime
from get_data import get_data_guigang, get_data_lengban, get_data_reban, \
    get_data_suanxiban, get_data_duxinban, get_data_diandusi, get_data_pensusi

# <editor-fold desc="参数设置">
date = datetime.now()
today = str(date.month) + '月' + str(date.day) + '日'
day = date.day
month = date.month
# </editor-fold>

# <editor-fold desc="采集数据并导出">
writer = pd.ExcelWriter('./每日明细数据/最新数据' + today + '.xlsx')
data_guigang = get_data_guigang()
name_guigang = list(data_guigang['表格表头'])
data_guigang.to_excel(writer, sheet_name='硅钢', index=False)

data_lengban = get_data_lengban()
name_lengban = list(data_lengban['表格表头'])
data_lengban.to_excel(writer, sheet_name='冷板', index=False)

data_reban = get_data_reban()
name_reban = list(data_reban['表格表头'])
data_reban.to_excel(writer, sheet_name='热板', index=False)

data_suanxiban = get_data_suanxiban()
name_suanxiban = list(data_suanxiban['表格表头'])
data_suanxiban.to_excel(writer, sheet_name='酸洗板', index=False)

data_duxinban = get_data_duxinban()
name_duxinban = list(data_duxinban['表格表头'])
data_duxinban.to_excel(writer, sheet_name='镀锌板', index=False)

data_diandusi = get_data_diandusi()
data_diandusi.to_excel(writer, sheet_name='电镀丝', index=False)

data_pensusi = get_data_pensusi()
data_pensusi.to_excel(writer, sheet_name='喷塑丝', index=False)

writer.save()
# </editor-fold>

# <editor-fold desc="读取数据表格">
# 读取待填充表格
file_path = '原始数据.xlsx'
wb = openpyxl.load_workbook(file_path)
# 读取采集的数据
new_data_file_path = './每日明细数据/最新数据' + today + '.xlsx'
wb_new = openpyxl.load_workbook(new_data_file_path)


# </editor-fold>


# <editor-fold desc="将数据填入表格">
def export_data_guigang():
    sh = wb['硅钢']
    sh_new = wb_new['硅钢']

    # 获取待填充单元格坐标
    def get_input_loc(day, month, name):
        for j_0 in range(1, 100, 1):
            if sh.cell(2, j_0).value == str(month) + '月份（吨）':
                col_0 = j_0 - 1
        for i_loc in range(1, 100, 1):
            if sh.cell(i_loc, 1).value == str(day) + '号':
                row_loc = i_loc
        for j_loc in range(1, 6, 1):
            if sh.cell(3, j_loc + col_0).value == name:
                col_loc = j_loc + col_0
        return row_loc, col_loc

    # 获取最新数据坐标
    def get_data_loc(name):
        for i_new in range(1, 10, 1):
            for j_new in range(1, 10, 1):
                if sh_new.cell(i_new, j_new).value == name:
                    data_row_loc = i_new
                    data_col_loc = j_new - 1
        return data_row_loc, data_col_loc

    for name in name_guigang:
        sh.cell(get_input_loc(day, month, name)[0], get_input_loc(day, month, name)[1]).value = int(
            sh_new.cell(get_data_loc(name)[0], get_data_loc(name)[1]).value)


def export_data_lengban():
    sh = wb['冷板']
    sh_new = wb_new['冷板']

    # 获取待填充单元格坐标
    def get_input_loc(day, month, name):
        for j_0 in range(1, 100, 1):
            if sh.cell(2, j_0).value == str(month) + '月份（吨）':
                col_0 = j_0 - 1
        for i_loc in range(1, 100, 1):
            if sh.cell(i_loc, 1).value == str(day) + '号':
                row_loc = i_loc
        for j_loc in range(1, 9, 1):
            if sh.cell(3, j_loc + col_0).value == name:
                col_loc = j_loc + col_0
        return row_loc, col_loc

    # 获取最新数据坐标
    def get_data_loc(name):
        for i_new in range(1, 10, 1):
            for j_new in range(1, 10, 1):
                if sh_new.cell(i_new, j_new).value == name:
                    data_row_loc = i_new
                    data_col_loc = j_new - 1
        return data_row_loc, data_col_loc

    for name in name_lengban:
        sh.cell(get_input_loc(day, month, name)[0], get_input_loc(day, month, name)[1]).value = int(
            sh_new.cell(get_data_loc(name)[0], get_data_loc(name)[1]).value)


def export_data_reban():
    sh = wb['热轧板卷']
    sh_new = wb_new['热板']

    # 获取待填充单元格坐标
    def get_input_loc(day, month, name):
        for j_0 in range(1, 100, 1):
            if sh.cell(2, j_0).value == str(month) + '月份（吨）':
                col_0 = j_0 - 1
        for i_loc in range(1, 100, 1):
            if sh.cell(i_loc, 1).value == str(day) + '号':
                row_loc = i_loc
        for j_loc in range(1, 6, 1):
            if sh.cell(3, j_loc + col_0).value == name:
                col_loc = j_loc + col_0
        return row_loc, col_loc

    # 获取最新数据坐标
    def get_data_loc(name):
        for i_new in range(1, 20, 1):
            for j_new in range(1, 20, 1):
                if sh_new.cell(i_new, j_new).value == name:
                    data_row_loc = i_new
                    data_col_loc = j_new - 1
        return data_row_loc, data_col_loc

    for name in name_reban:
        sh.cell(get_input_loc(day, month, name)[0], get_input_loc(day, month, name)[1]).value = int(
            sh_new.cell(get_data_loc(name)[0], get_data_loc(name)[1]).value)


def export_data_suanxiban():
    sh = wb['酸洗板']
    sh_new = wb_new['酸洗板']

    # 获取待填充单元格坐标
    def get_input_loc(day, month, name):
        for j_0 in range(1, 100, 1):
            # 酸洗板月份在第一行
            if sh.cell(1, j_0).value == str(month) + '月份':
                col_0 = j_0 - 1
        for i_loc in range(1, 100, 1):
            if sh.cell(i_loc, 1).value == str(day) + '号':
                row_loc = i_loc
        for j_loc in range(1, 3, 1):
            if sh.cell(2, j_loc + col_0).value == name:
                col_loc = j_loc + col_0
        return row_loc, col_loc

    # 获取最新数据坐标
    def get_data_loc(name):
        for i_new in range(1, 10, 1):
            for j_new in range(1, 10, 1):
                if sh_new.cell(i_new, j_new).value == name:
                    data_row_loc = i_new
                    data_col_loc = j_new - 1
        return data_row_loc, data_col_loc

    for name in name_suanxiban:
        sh.cell(get_input_loc(day, month, name)[0], get_input_loc(day, month, name)[1]).value = int(
            sh_new.cell(get_data_loc(name)[0], get_data_loc(name)[1]).value)


def export_data_duxinban():
    sh = wb['镀锌板']
    sh_new = wb_new['镀锌板']

    # 获取待填充单元格坐标
    def get_input_loc(day, month, name):
        for j_0 in range(1, 100, 1):
            # 酸洗板月份在第一行
            if sh.cell(2, j_0).value == str(month) + '月份（吨）':
                col_0 = j_0 - 1
        for i_loc in range(1, 100, 1):
            if sh.cell(i_loc, 1).value == str(day) + '号':
                row_loc = i_loc
        # 仅需采集4个值
        for j_loc in range(1, 5, 1):
            if sh.cell(3, j_loc + col_0).value == name:
                col_loc = j_loc + col_0
        return row_loc, col_loc

    # 获取最新数据坐标
    def get_data_loc(name):
        for i_new in range(1, 20, 1):
            for j_new in range(1, 20, 1):
                if sh_new.cell(i_new, j_new).value == name:
                    data_row_loc = i_new
                    data_col_loc = j_new - 1
        return data_row_loc, data_col_loc

    for name in name_duxinban:
        sh.cell(get_input_loc(day, month, name)[0], get_input_loc(day, month, name)[1]).value = int(
            sh_new.cell(get_data_loc(name)[0], get_data_loc(name)[1]).value)


def export_data_diandusi():
    sh = wb['电镀丝']
    sh_new = wb_new['电镀丝']

    # 获取待填充单元格坐标
    def get_input_loc(day, month):
        for j_loc in range(1, 100, 1):
            if sh.cell(2, j_loc).value == str(month) + '月份':
                col_loc = j_loc
        for i_loc in range(1, 100, 1):
            if sh.cell(i_loc, 1).value == str(day) + '号':
                row_loc = i_loc
        return row_loc, col_loc

    sh.cell(get_input_loc(day, month)[0], get_input_loc(day, month)[1]).value = int(sh_new.cell(2, 5).value)


def export_data_pensusi():
    sh = wb['喷塑丝']
    sh_new = wb_new['喷塑丝']

    # 获取待填充单元格坐标
    def get_input_loc(day, month):
        for j_loc in range(1, 100, 1):
            if sh.cell(2, j_loc).value == str(month) + '月份':
                col_loc = j_loc
        for i_loc in range(1, 100, 1):
            if sh.cell(i_loc, 1).value == str(day) + '号':
                row_loc = i_loc
        return row_loc, col_loc

    sh.cell(get_input_loc(day, month)[0], get_input_loc(day, month)[1]).value = int(sh_new.cell(2, 5).value)


export_data_guigang()
export_data_lengban()
export_data_reban()
export_data_suanxiban()
export_data_duxinban()
export_data_diandusi()
export_data_pensusi()
# </editor-fold>
wb.save('./数据更新/' + today + '更新数据.xlsx')

