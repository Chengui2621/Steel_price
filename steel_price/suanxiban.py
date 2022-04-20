import requests
from lxml import etree
import re
import pandas as pd
import openpyxl
from datetime import datetime

url = 'https://list1.mysteel.com/market/p-231-----010103-0-010101-------1.html'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36',
    'Cookie': '_last_loginuname=goodbaby123; _login_psd=3ba7e75e56a34c01b56c65f10e8901b16; _rememberStatus=true; Hm_lvt_1c4432afacfa2301369a5625795031b8=1649222144; allHistory=%5B%22%E7%94%B5%E9%95%80%E4%B8%9D%22%5D; href=http%3A%2F%2Fguigang.mysteel.com%2F; accessId=5d36a9e0-919c-11e9-903c-ab24dbab411b; pageViewNum=3; _last_ch_r_t=1649397611208; fingerprint=a64ff2e9c89fde087b97042290980c0c; _login_token=3b4df51e8f109f30b5166808f3ce710b; _login_uid=2243467; _login_mid=3057897; _login_ip=183.129.214.218; 3b4df51e8f109f30b5166808f3ce710b=1%3D10%262%3D10%26catalog%3D010102%2C010103; Hm_lpvt_1c4432afacfa2301369a5625795031b8=1649398288'}

response = requests.get(url)
url_tree = etree.HTML(response.text)
url_list = url_tree.xpath('//*[@id="articleList"]/ul/li')
first_title = url_list[0].xpath('./a/text()')[0]
title_check = re.search('([0-9]+?月[0-9]+?日)', first_title).group()
for li in url_list:
    next_title = li.xpath('./a/text()')
    if next_title:
        if '热轧酸洗卷' in next_title[0] and '（' not in next_title[0] and '(' not in next_title[0] and title_check in next_title[0]:
            print(next_title[0])
            next_url = li.xpath('./a/@href')[0]

detail_response = requests.get(next_url, headers=headers)
detail_url_tree = etree.HTML(detail_response.text)
# 获取数据标题和日期
detail_title = detail_url_tree.xpath('//*[@id="articleContent"]/h1/text()')[0]
detail_data_date = re.search('([0-9]+?月[0-9]+?日)', detail_title).group()
detail_data_month = detail_data_date.split('月')[0] + '月份（吨）'
detail_data_day = detail_data_date.split('月')[1][0] + '号'
detail_tr_list = detail_url_tree.xpath('//*[@id="marketTable"]/tr')
detail_tr_list = [detail_tr_list[21], detail_tr_list[4]]


data_list = []
for tr in detail_tr_list:
    td_list = tr.xpath('./td[@align]')
    td_list = [td_list[0], td_list[1], td_list[2], td_list[3], td_list[4]]
    single_data_list = []
    for td in td_list:
        data = td.xpath('./text()')[0].replace('\r', '').replace('\n', '').replace('\t', '').replace(' ', '')
        single_data_list.append(data)
    data_list.append(single_data_list)
df = pd.DataFrame(data_list)
table_title = ['4.0酸洗板', '2.0酸洗板']
df.columns = ['品名', '规格', '材质', '钢厂', '价格']
df['表格表头'] = table_title
print(df)
df.to_excel('./每日明细数据/最新数据' + detail_data_date + '.xlsx', sheet_name='酸洗板', index=False)

name_list = table_title

date = datetime.now()
today = str(date.month) + '月' + str(date.day) + '日'
day = date.day
month = date.month
file_path = '原始数据.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.sheetnames
sh = wb['酸洗板']

new_data_file_path = './每日明细数据/最新数据' + today + '.xlsx'
wb_new = openpyxl.load_workbook(new_data_file_path)
sheet_new = wb_new.sheetnames
sh_new = wb_new['酸洗板']


# 获取待填充单元格坐标
def get_input_loc(day, month, name):
    for j_0 in range(1, 100, 1):
        # 酸洗板月份在第一行
        if sh.cell(1, j_0).value == str(month) + '月份':
            col_0 = j_0 - 1
    for i_loc in range(1, 100, 1):
        if sh.cell(i_loc, 1).value == str(day) + '号':
            row_loc = i_loc
    for j_loc in range(1, 3, 1):
        if sh.cell(2, j_loc + col_0).value == name:
            col_loc = j_loc + col_0
    return row_loc, col_loc


# 获取最新数据坐标
def get_data_loc(name):
    for i_new in range(1, 20, 1):
        for j_new in range(1, 20, 1):
            if sh_new.cell(i_new, j_new).value == name:
                data_row_loc = i_new
                data_col_loc = j_new - 1
    return data_row_loc, data_col_loc


for name in name_list:
    sh.cell(get_input_loc(day, month, name)[0], get_input_loc(day, month, name)[1]).value = int(
        sh_new.cell(get_data_loc(name)[0], get_data_loc(name)[1]).value)

wb.save('./数据更新/' + today + '更新数据.xlsx')
